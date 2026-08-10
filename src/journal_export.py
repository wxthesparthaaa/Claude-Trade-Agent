"""
Builds an Excel workbook from the trade journal for offline review --
generated on demand from trade_journal.json (the source of truth,
GitHub-synced), not maintained as a separate persisted binary file that
could drift out of sync. Direct port of the sibling Forex Agent
project's journal_export.py, columns adjusted to this project's
JournalEntry shape (no stop-loss/take-profit/R-multiple fields -- this
project's exits are systematic rule re-checks every scan, not fixed
levels attached at entry).
"""
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

COLUMNS = [
    "Symbol", "Sleeve", "Type", "Quantity", "Entry Price", "Confidence %",
    "Status", "Opened At", "Closed At", "Exit Price", "Realized P&L", "Reason",
]


def build_journal_workbook(entries: list) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Trade Journal"

    ws.append(COLUMNS)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for entry in entries:
        ws.append([
            entry.get("symbol"), entry.get("sleeve"), entry.get("position_type"), entry.get("quantity"),
            entry.get("entry_price"), entry.get("confidence_pct"), entry.get("status"),
            entry.get("opened_at"), entry.get("closed_at"), entry.get("exit_price"),
            entry.get("realized_pnl"), entry.get("reason"),
        ])

    for i, _ in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = 16
    ws.column_dimensions[get_column_letter(len(COLUMNS))].width = 60  # Reason

    return wb
